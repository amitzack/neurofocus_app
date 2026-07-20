## -*- coding: utf-8 -*-
"""
EEG Analysis - ADHD vs Control
גישה: Windowed (כל חלון = דוגמה נפרדת)
פיצרים: 1,501 לחלון (817 ערוץ-בודד + 684 coherence)
חלוקה: רנדומלית של חלונות (ללא קשר לנבדקים)
בחירת פיצרים: ACC / MRMR / COMBINED (בחר ב-FEATURE_SELECTION_MODE)
"""

import numpy as np
import os
import sys
import pandas as pd
import datetime
import scipy.stats
from joblib import Parallel, delayed, dump

# תיקון לקונסולה של Windows (cp1255/cp1252) שלא תומכת בתווים כמו -> — +/-
try:
    sys.stdout.reconfigure(encoding="utf-8")
    sys.stderr.reconfigure(encoding="utf-8")
except Exception:
    pass
import json

from sklearn.model_selection import train_test_split, cross_val_score, StratifiedKFold
from sklearn.metrics import confusion_matrix, accuracy_score, f1_score, roc_auc_score
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from sklearn.neighbors import KNeighborsClassifier

# =======================================================
# FLAGS
# =======================================================
RUN_MODEL_TOURNAMENT   = False
RUN_FINAL_TEST         = False

# ← זה החלק החדש היחיד: שומר את המודל הסופי לקבצים, כדי שדף
#   הנחיתה יוכל לטעון אותו ולהריץ תחזיות אמיתיות.
SAVE_FINAL_MODEL   = True
FINAL_N_FEATURES   = 240          # כמה פיצרים במודל שנשמר (כמו בעבודה הסופית)
MODEL_EXPORT_DIR   = "model_export"

# ← שנה כאן את שיטת הדירוג: "ACC" / "MRMR" / "COMBINED"
FEATURE_SELECTION_MODE = "ACC"

# =======================================================
# PATHS
# =======================================================
FOLDER         = r"C:\Users\משתמש\Downloads\features_new"
EXCEL_OUT_PATH = os.path.join(FOLDER, "Results_Windowed_Coherence.xlsx")

# נתיבי cache לדירוגים
ACC_CACHE     = os.path.join(FOLDER, "windowed_acc_ranking.npy")
MRMR_CACHE    = os.path.join(FOLDER, "windowed_mrmr_ranking.npy")
COMBINED_CACHE= os.path.join(FOLDER, "windowed_combined_ranking.npy")

# נתיבי cache לציונים (במקביל לדירוגים)
ACC_SCORES_CACHE      = os.path.join(FOLDER, "windowed_acc_scores.npy")
MRMR_SCORES_CACHE     = os.path.join(FOLDER, "windowed_mrmr_scores.npy")
COMBINED_SCORES_CACHE = os.path.join(FOLDER, "windowed_combined_scores.npy")

# =======================================================
# PARAMETERS
# =======================================================
FEAT_START   = 100       # ← פיצר התחלה
FEAT_END     = 155     # ← פיצר סיום (None = עד הסוף)
STEP         = 5       # ← קפיצות
TEST_SIZE    = 0.2
N_FOLDS      = 5
APPLY_ZSCORE = True
N_JOBS       = -1

ADHD_IDS    = list(range(1, 62))
CONTROL_IDS = list(range(62, 122))

# =======================================================
# MODELS
# =======================================================
MODELS = [
    {
        "name":   "KNN",
        "class":  KNeighborsClassifier,
        "params": {"n_neighbors": 3, "weights": "distance", "metric": "euclidean"},
    },
]

# =======================================================
# FEATURE SELECTION FUNCTIONS
# =======================================================

def acc_score_single(feature_values, labels):
    """ACC scoring לפיצר אחד — סף אופטימלי"""
    best_acc = 0.0
    for th in np.unique(feature_values):
        acc     = np.mean((feature_values >  th).astype(int) == labels)
        acc_inv = np.mean((feature_values <= th).astype(int) == labels)
        best_acc = max(best_acc, acc, acc_inv)
    return best_acc


def compute_acc_ranking(X_train, y_train, cache_path, scores_cache_path=None):
    """מחשב ACC ranking על X_train ושומר cache (דירוג + ציון)"""
    if scores_cache_path is None:
        scores_cache_path = cache_path.replace("_ranking.npy", "_scores.npy")

    if os.path.exists(cache_path) and os.path.exists(scores_cache_path):
        print(f"  Loading ACC ranking + scores from cache: {cache_path}")
        return np.load(cache_path).tolist()

    print("  Computing ACC ranking (parallel)...")
    scores = Parallel(n_jobs=N_JOBS)(
        delayed(acc_score_single)(X_train[:, i], y_train)
        for i in range(X_train.shape[1])
    )
    scores  = np.array(scores)
    ranking = np.argsort(scores)[::-1].tolist()
    ranked_scores = scores[ranking]                    # הציון של כל פיצר, באותו סדר כמו הדירוג
    np.save(cache_path, np.array(ranking))
    np.save(scores_cache_path, ranked_scores)
    print(f"  ACC ranking saved:  {cache_path}")
    print(f"  ACC scores saved:   {scores_cache_path}")
    return ranking


def _mrmr_selection_scores(selected, relevance, redundancy):
    """
    מחשב את ציון ה-mRMR של כל פיצר בצעד שבו נבחר.
    ציון = relevance / mean(redundancy מול הפיצרים שכבר נבחרו).
    הפיצר הראשון: ציון = relevance בלבד (אין redundancy עדיין).
    """
    scores = []
    for step, feat in enumerate(selected):
        rel = float(relevance.loc[feat])
        if step == 0:
            scores.append(rel)
        else:
            prev = selected[:step]
            red_vals = [abs(float(redundancy.loc[feat, p])) for p in prev]
            red_mean = np.mean(red_vals) if red_vals else 1.0
            if red_mean == 0:
                red_mean = 1e-9
            scores.append(rel / red_mean)
    return np.array(scores, dtype=float)


def compute_mrmr_ranking(X_train, y_train, cache_path, n_features, scores_cache_path=None):
    """מחשב mRMR ranking ושומר cache (דירוג + ציון)"""
    if scores_cache_path is None:
        scores_cache_path = cache_path.replace("_ranking.npy", "_scores.npy")

    if os.path.exists(cache_path) and os.path.exists(scores_cache_path):
        print(f"  Loading MRMR ranking + scores from cache: {cache_path}")
        return np.load(cache_path).tolist()

    print("  Computing MRMR ranking (this may take a while)...")
    from mrmr import mrmr_classif
    X_df     = pd.DataFrame(X_train)
    y_series = pd.Series(y_train)
    selected, relevance, redundancy = mrmr_classif(
        X=X_df, y=y_series, K=n_features,
        return_scores=True, show_progress=True
    )
    ranking = [int(c) for c in selected]
    scores  = _mrmr_selection_scores(selected, relevance, redundancy)
    np.save(cache_path, np.array(ranking))
    np.save(scores_cache_path, scores)
    print(f"  MRMR ranking saved: {cache_path}")
    print(f"  MRMR scores saved:  {scores_cache_path}")
    return ranking


def compute_combined_ranking(X_train, y_train, acc_cache, combined_cache, n_top_acc=250,
                             scores_cache_path=None):
    """
    COMBINED: ACC על כל הפיצרים -> top-N -> mRMR על top-N
    """
    if scores_cache_path is None:
        scores_cache_path = combined_cache.replace("_ranking.npy", "_scores.npy")

    if os.path.exists(combined_cache) and os.path.exists(scores_cache_path):
        print(f"  Loading COMBINED ranking + scores from cache: {combined_cache}")
        return np.load(combined_cache).tolist()

    # שלב 1: ACC
    acc_ranking = compute_acc_ranking(X_train, y_train, acc_cache)
    top200_idx  = acc_ranking[:n_top_acc]

    # שלב 2: mRMR על top-N
    print(f"  Running mRMR on top-{n_top_acc} ACC features...")
    from mrmr import mrmr_classif
    X_sub    = X_train[:, top200_idx]
    X_df     = pd.DataFrame(X_sub)
    y_series = pd.Series(y_train)
    local_selected, relevance, redundancy = mrmr_classif(
        X=X_df, y=y_series, K=n_top_acc,
        return_scores=True, show_progress=True
    )
    local_ranking = [int(c) for c in local_selected]

    # ציון mRMR של כל פיצר בצעד שבו נבחר (באינדוקס המקומי של top-N)
    scores = _mrmr_selection_scores(local_selected, relevance, redundancy)

    # המרה בחזרה לאינדקסי פיצרים מקוריים
    final_ranking = [top200_idx[i] for i in local_ranking]
    np.save(combined_cache, np.array(final_ranking))
    np.save(scores_cache_path, scores)
    print(f"  COMBINED ranking saved: {combined_cache}")
    print(f"  COMBINED scores saved:  {scores_cache_path}")
    return final_ranking


def get_ranking(X_train, y_train, mode):
    """מחזיר רשימת אינדקסים ממוינת לפי שיטת הדירוג שנבחרה"""
    if mode == "ACC":
        return compute_acc_ranking(X_train, y_train, ACC_CACHE)
    elif mode == "MRMR":
        return compute_mrmr_ranking(X_train, y_train, MRMR_CACHE, X_train.shape[1])
    elif mode == "COMBINED":
        return compute_combined_ranking(X_train, y_train, ACC_CACHE, COMBINED_CACHE)
    else:
        raise ValueError(f"FEATURE_SELECTION_MODE לא מוכר: {mode}. בחר ACC / MRMR / COMBINED")

# =======================================================
# EXCEL LOGGING
# =======================================================

def save_to_excel(cv_log, test_log, meta, path):
    hdr_fill = PatternFill("solid", fgColor="1F4E79")
    hdr_font = Font(bold=True, color="FFFFFF")
    par_fill = PatternFill("solid", fgColor="D9E1F2")
    bst_fill = PatternFill("solid", fgColor="E2EFDA")
    center   = Alignment(horizontal="center")
    border   = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"),  bottom=Side(style="thin")
    )

    skip_params = {"random_state", "verbose", "n_jobs", "thread_count",
                   "nthread", "eval_metric"}

    def params_str(m):
        return ", ".join(f"{k}={v}" for k, v in m["params"].items()
                         if k not in skip_params)

    def style(cell, fill=None):
        cell.alignment = center
        cell.border    = border
        if fill:
            cell.fill = fill

    exists = os.path.exists(path)
    wb = load_workbook(path) if exists else Workbook()
    if not exists and "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    meta_keys = ["Timestamp", "Selection Mode", "N Features",
                 "Test Size %", "N Folds", "Train Windows", "Test Windows",
                 "Classifier Params"]
    meta_vals = [
        meta["timestamp"], meta["selection_mode"], meta["n_features"],
        meta["test_size_pct"], meta["n_folds"],
        meta["train_samples"], meta["test_samples"], ""
    ]

    # --- CV sheet ---
    cv_headers = meta_keys + ["Model", "CV Mean", "CV Std", "Test Accuracy"]
    if "CV Results" not in wb.sheetnames:
        ws1 = wb.create_sheet("CV Results")
        for col, h in enumerate(cv_headers, 1):
            c = ws1.cell(row=1, column=col, value=h)
            c.fill, c.font, c.alignment, c.border = hdr_fill, hdr_font, center, border
    else:
        ws1 = wb["CV Results"]

    test_acc_map = {r["Model"]: r["Accuracy"] for r in test_log}
    best_cv = max(r["CV Mean"] for r in cv_log) if cv_log else None
    row = ws1.max_row + 1
    for r in cv_log:
        fill = bst_fill if r["CV Mean"] == best_cv else None
        meta_vals[-1] = r.get("params_str", "")
        vals = meta_vals + [r["Model"], round(r["CV Mean"], 4),
                            round(r["CV Std"], 4),
                            test_acc_map.get(r["Model"], "")]
        for col, val in enumerate(vals, 1):
            style(ws1.cell(row=row, column=col, value=val),
                  fill if col > len(meta_vals) else par_fill)
        row += 1

    # --- Test sheet ---
    test_headers = meta_keys + ["Model", "Accuracy", "F1", "AUC",
                                 "Sensitivity", "Specificity",
                                 "TP", "FP", "FN", "TN"]
    if "Test Results" not in wb.sheetnames:
        ws2 = wb.create_sheet("Test Results")
        for col, h in enumerate(test_headers, 1):
            c = ws2.cell(row=1, column=col, value=h)
            c.fill, c.font, c.alignment, c.border = hdr_fill, hdr_font, center, border
    else:
        ws2 = wb["Test Results"]

    best_test = max(r["Accuracy"] for r in test_log) if test_log else None
    row = ws2.max_row + 1
    for r in test_log:
        fill = bst_fill if r["Accuracy"] == best_test else None
        meta_vals[-1] = r.get("params_str", "")
        tp, fp, fn, tn = r["TP"], r["FP"], r["FN"], r["TN"]
        sens = tp / (tp + fn) if (tp + fn) > 0 else 0
        spec = tn / (tn + fp) if (tn + fp) > 0 else 0
        vals = meta_vals + [
            r["Model"], round(r["Accuracy"], 4), round(r["F1"], 4),
            round(r["AUC"], 4), round(sens, 4), round(spec, 4),
            tp, fp, fn, tn
        ]
        for col, val in enumerate(vals, 1):
            style(ws2.cell(row=row, column=col, value=val),
                  fill if col > len(meta_vals) else par_fill)
        row += 1

    wb.save(path)
    print(f"  Excel updated: {path}")

# =======================================================
# MAIN
# =======================================================

if __name__ == "__main__":

    # -------------------------------------------------------
    # שלב 1: טעינת חלונות מכל הנבדקים
    # -------------------------------------------------------
    print("=" * 60)
    print("Loading windowed features from all subjects...")
    print("=" * 60)

    X_list, y_list = [], []
    for i in range(1, 122):
        fp = os.path.join(FOLDER, f"features_subj{i}.npy")
        lp = os.path.join(FOLDER, f"labels_subj{i}.npy")
        if not os.path.exists(fp):
            continue
        windows = np.load(fp)   # (n_windows, 1501)
        labels  = np.load(lp)   # (n_windows,)
        X_list.append(windows)
        y_list.append(labels)
        print(f"  Subject {i}: {windows.shape[0]} windows")

    X_all = np.vstack(X_list)
    y_all = np.concatenate(y_list)
    print(f"\nTotal: {X_all.shape} | ADHD windows={np.sum(y_all==0)} | Control windows={np.sum(y_all==1)}")

    # -------------------------------------------------------
    # שלב 2: ניקוי NaN / Inf
    # -------------------------------------------------------
    print("\nCleaning NaN/Inf...")
    X_all[np.isinf(X_all)] = np.nan
    for col in range(X_all.shape[1]):
        nan_mask = np.isnan(X_all[:, col])
        if nan_mask.any():
            col_mean = np.nanmean(X_all[:, col])
            X_all[nan_mask, col] = col_mean if not np.isnan(col_mean) else 0.0
    print("  Done.")

    # -------------------------------------------------------
    # שלב 3: חלוקה רנדומלית של חלונות (ללא קשר לנבדקים)
    # -------------------------------------------------------
    X_train, X_test, y_train, y_test = train_test_split(
        X_all, y_all,
        test_size=TEST_SIZE,
        random_state=42,
        stratify=y_all
    )
    print(f"\nTrain: {len(y_train)} windows | Test: {len(y_test)} windows")
    print(f"Train ADHD={np.sum(y_train==0)} | Control={np.sum(y_train==1)}")
    print(f"Test  ADHD={np.sum(y_test==0)}  | Control={np.sum(y_test==1)}")

    # -------------------------------------------------------
    # שלב 4: דירוג פיצרים (על train בלבד — פעם אחת)
    # -------------------------------------------------------
    print(f"\n--- FEATURE RANKING: {FEATURE_SELECTION_MODE} ---")
    ranking = get_ranking(X_train, y_train, FEATURE_SELECTION_MODE)

    max_features   = len(ranking)
    feat_end       = min(FEAT_END, max_features) if FEAT_END is not None else max_features
    feature_counts = list(range(FEAT_START, feat_end + 1, STEP))
    if not feature_counts or feature_counts[-1] != feat_end:
        feature_counts.append(feat_end)

    print(f"  Loop: {FEAT_START} -> {feat_end} (step={STEP}) | {len(feature_counts)} runs total")

    # -------------------------------------------------------
    # שמירת המודל הסופי (לשימוש דף הנחיתה) — לא משנה שום דבר
    # אחר, רק מוסיף את זה כאן, לפני שאר הלולאה/היציאה.
    # -------------------------------------------------------
    if SAVE_FINAL_MODEL:
        print(f"\n--- SAVING FINAL MODEL ({FINAL_N_FEATURES} features) ---")
        final_idx = ranking[:FINAL_N_FEATURES]
        X_train_final = X_train[:, final_idx]
        X_test_final  = X_test[:,  final_idx]

        mean = X_train_final.mean(axis=0)
        std  = X_train_final.std(axis=0)
        std[std == 0] = 1
        X_train_final_norm = (X_train_final - mean) / std
        X_test_final_norm  = (X_test_final  - mean) / std

        final_clf = KNeighborsClassifier(**MODELS[0]["params"])
        final_clf.fit(X_train_final_norm, y_train)

        y_pred_final = final_clf.predict(X_test_final_norm)
        final_acc = accuracy_score(y_test, y_pred_final)
        final_f1  = f1_score(y_test, y_pred_final, zero_division=0)
        print(f"  Sanity check -> Accuracy={final_acc:.4f} | F1={final_f1:.4f}")

        os.makedirs(MODEL_EXPORT_DIR, exist_ok=True)
        dump(final_clf, os.path.join(MODEL_EXPORT_DIR, "knn_model.joblib"))
        np.save(os.path.join(MODEL_EXPORT_DIR, "selected_indices.npy"), np.array(final_idx))
        np.save(os.path.join(MODEL_EXPORT_DIR, "feat_mean.npy"), mean)
        np.save(os.path.join(MODEL_EXPORT_DIR, "feat_std.npy"), std)

        meta = {
            "raw_feature_len": int(X_all.shape[1]),
            "n_features_selected": int(FINAL_N_FEATURES),
            "feature_selection_mode": FEATURE_SELECTION_MODE,
            "knn_params": MODELS[0]["params"],
            "label_map": {"0": "ADHD", "1": "Control"},
            "sanity_check_accuracy": float(final_acc),
            "sanity_check_f1": float(final_f1),
        }
        with open(os.path.join(MODEL_EXPORT_DIR, "meta.json"), "w", encoding="utf-8") as f:
            json.dump(meta, f, indent=2, ensure_ascii=False)

        print(f"  Saved 5 files to ./{MODEL_EXPORT_DIR}/ :")
        for fn in ["knn_model.joblib", "selected_indices.npy", "feat_mean.npy",
                   "feat_std.npy", "meta.json"]:
            print(f"    - {fn}")


    # -------------------------------------------------------
    # הגדרות קבועות לכל הלולאה
    # -------------------------------------------------------
    skf = StratifiedKFold(n_splits=N_FOLDS, shuffle=True, random_state=42)

    skip_keys = {"random_state", "verbose", "n_jobs", "thread_count",
                 "nthread", "eval_metric"}

    def params_str(m):
        return ", ".join(f"{k}={v}" for k, v in m["params"].items()
                         if k not in skip_keys)

    # -------------------------------------------------------
    # שלב 5: לולאה על מספרי פיצרים
    # רצה רק אם מבקשים סיווג. אם רק רוצים דירוג+ציון — מדלגים.
    # -------------------------------------------------------
    if not (RUN_MODEL_TOURNAMENT or RUN_FINAL_TEST):
        print("\nTournament & Final test are OFF - ranking + scores saved, skipping classification.")
        print("\n=== DONE ===")
        import sys as _sys
        _sys.exit(0)

    for n_features in feature_counts:
        print(f"\n{'='*60}")
        print(f"  n_features = {n_features}")
        print(f"{'='*60}")

        selected_idx = ranking[:n_features]
        X_train_sel  = X_train[:, selected_idx]
        X_test_sel   = X_test[:,  selected_idx]

        # נרמול Z-Score על בסיס train בלבד
        if APPLY_ZSCORE:
            mean = X_train_sel.mean(axis=0)
            std  = X_train_sel.std(axis=0)
            std[std == 0] = 1
            X_train_sel = (X_train_sel - mean) / std
            X_test_sel  = (X_test_sel  - mean) / std

        ts       = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        cv_log   = []
        test_log = []
        if RUN_MODEL_TOURNAMENT:
            print(f"\n--- CROSS VALIDATION ({N_FOLDS}-Fold) ---")

            def run_cv(m):
                scores = cross_val_score(
                    m["class"](**m["params"]), X_train_sel, y_train,
                    cv=skf, n_jobs=1
                )
                return {
                    "Model":      m["name"],
                    "CV Mean":    float(np.mean(scores)),
                    "CV Std":     float(np.std(scores)),
                    "params_str": params_str(m)
                }

            cv_log = Parallel(n_jobs=N_JOBS)(delayed(run_cv)(m) for m in MODELS)
            for r in cv_log:
                print(f"  {r['Model']:16s} | CV = {r['CV Mean']:.4f} +/- {r['CV Std']:.4f}")

        if RUN_FINAL_TEST:
            print("\n--- FINAL TEST ---")
            for m in MODELS:
                clf    = m["class"](**m["params"])
                clf.fit(X_train_sel, y_train)
                y_pred = clf.predict(X_test_sel)
                acc    = accuracy_score(y_test, y_pred)
                f1     = f1_score(y_test, y_pred, zero_division=0)
                try:    auc = roc_auc_score(y_test, clf.predict_proba(X_test_sel)[:, 1])
                except: auc = np.nan
                tn, fp, fn, tp = confusion_matrix(y_test, y_pred).ravel()
                cv_mean = next((r["CV Mean"] for r in cv_log if r["Model"] == m["name"]), np.nan)
                ps      = next((r["params_str"] for r in cv_log if r["Model"] == m["name"]),
                               params_str(m))
                test_log.append({
                    "Model":      m["name"],
                    "Accuracy":   float(acc),
                    "F1":         float(f1),
                    "AUC":        float(auc),
                    "TP": int(tp), "FP": int(fp), "FN": int(fn), "TN": int(tn),
                    "params_str": ps
                })
                print(f"  {m['name']:16s} | CV={cv_mean:.4f} | ACC={acc:.4f} | F1={f1:.4f}")

        # שמירה ל-Excel אחרי כל n_features
        meta = {
            "timestamp":      ts,
            "selection_mode": FEATURE_SELECTION_MODE,
            "n_features":     n_features,
            "test_size_pct":  int(TEST_SIZE * 100),
            "n_folds":        N_FOLDS,
            "train_samples":  len(y_train),
            "test_samples":   len(y_test),
        }
        save_to_excel(cv_log, test_log, meta, EXCEL_OUT_PATH)

    print("\n=== DONE ===")