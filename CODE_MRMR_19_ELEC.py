# -*- coding: utf-8 -*-
"""
EEG Analysis - ADHD vs Control
ADHD=0, Control=1
"""

import numpy as np
import os
import datetime
import scipy.stats
from scipy.io import loadmat
from scipy.signal import butter, filtfilt, welch, coherence
from fooof import FOOOF
import antropy as ant
import pywt
from joblib import Parallel, delayed

from sklearn.feature_selection import mutual_info_classif, mutual_info_regression
from sklearn.model_selection import cross_val_score, StratifiedKFold
from sklearn.metrics import confusion_matrix, accuracy_score, f1_score, roc_auc_score
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from sklearn.neighbors import KNeighborsClassifier

# =======================================================
# FLAGS
# =======================================================
RUN_PREPROCESSING      = True
RUN_BUILD_DATASET      = True
ANALYSIS_MODE          = "statistics"
FEATURE_SELECTION_MODE = "MRMR"
RUN_MODEL_TOURNAMENT   = True
RUN_FINAL_TEST         = True

# =======================================================
# PARAMETERS
# =======================================================
fs     = 128
N_JOBS = -1

save_folder = r"C:\Users\משתמש\Downloads\features_new"

excel_log_path = os.path.join(save_folder, "Models_Results_19_elec.xlsx")

ADHD_IDS    = list(range(1, 62))
CONTROL_IDS = list(range(62, 122))

folders = {
    r'C:\Users\משתמש\Downloads\ADHD 1-61\ADHD 1-61': ADHD_IDS,
    r'C:\Users\משתמש\Downloads\Control 62-121\Control 62-121': CONTROL_IDS,
}

prms = {
    "num_channels":       19,
    "n_selected":         200,
    "ftrs_stats":         ["mean", "std", "median", "iqr", "skew", "kurtosis"],
    "lpf_cutoff":         30,
    "lpf_order":          4,
    "hpf_cutoff":         0.5,
    "hpf_order":          2,
    "apply_zscore":       True,
    "apply_final_zscore": True,
    "window_sec":         2,
    "overlap_pct":        0.5,
    "test_size_pct":      0.2,
    "n_folds":            5,   # ← שנה ידנית ל-3 / 4 / 5 לפני כל ריצה
}

# =======================================================
# MODELS
# =======================================================
MODELS_TO_COMPARE = [
    {
        "class":  KNeighborsClassifier,
        "params": {"n_neighbors": 5, "weights": "distance", "metric": "euclidean"},
        "name":   "KNN",
    },
]

os.makedirs(save_folder, exist_ok=True)

# =======================================================
# FILTERS & NORMALIZATION
# =======================================================
def lowpass_filter(signal):
    nyq = 0.5 * fs
    b, a = butter(prms["lpf_order"], prms["lpf_cutoff"] / nyq, btype='low')
    return filtfilt(b, a, signal, axis=0)

def highpass_filter(signal):
    nyq = 0.5 * fs
    b, a = butter(prms["hpf_order"], prms["hpf_cutoff"] / nyq, btype='high')
    return filtfilt(b, a, signal, axis=0)

def zscore_normalization(signal):
    std = np.std(signal, axis=0)
    std[std == 0] = 1
    return (signal - np.mean(signal, axis=0)) / std

# =======================================================
# WINDOWING
# =======================================================
def segment_signal(data):
    win_len = int(prms["window_sec"] * fs)
    step    = int(win_len * (1 - prms["overlap_pct"]))
    return np.array([data[i:i+win_len, :] for i in range(0, data.shape[0] - win_len + 1, step)])

# =======================================================
# FEATURE EXTRACTION
# =======================================================
def band_power(freqs, psd, fmin, fmax):
    idx = (freqs >= fmin) & (freqs <= fmax)
    return np.trapz(psd[idx], freqs[idx]) if np.any(idx) else 0

def time_features(window):
    feats = []
    for ch in range(window.shape[1]):
        x     = window[:, ch]
        diff1 = np.diff(x)
        diff2 = np.diff(diff1)
        act   = np.var(x)
        mob   = np.sqrt(np.var(diff1) / act) if act > 0 else 0
        comp  = np.sqrt(np.var(diff2) / np.var(diff1)) / mob if mob > 0 else 0
        feats.extend([
            np.mean(x), np.median(x), np.var(x, ddof=1), np.sqrt(np.mean(x**2)),
            np.ptp(x), scipy.stats.iqr(x), scipy.stats.skew(x), scipy.stats.kurtosis(x),
            np.sum(np.diff(np.sign(x)) != 0), np.sum(np.diff(np.sign(diff1)) != 0),
            act, mob, comp,
        ])
        entropy_fns = [ant.sample_entropy, ant.perm_entropy, ant.app_entropy,
                       getattr(ant, "fuzzy_entropy", None)]
        for fn in entropy_fns:
            if fn is None:
                feats.append(0)
                continue
            try:    feats.append(fn(x) if fn != ant.perm_entropy else fn(x, normalize=True))
            except: feats.append(0)
    return np.array(feats)

def frequency_power_features(window):
    feats = []
    for ch in range(window.shape[1]):
        freqs, psd = welch(window[:, ch], fs=fs, nperseg=fs)
        d, t, a, b = (band_power(freqs, psd, 1, 4), band_power(freqs, psd, 4, 8),
                      band_power(freqs, psd, 8, 13), band_power(freqs, psd, 13, 30))
        total = d + t + a + b + 1e-12
        feats.extend([d, t, a, b, d/total, t/total, a/total, b/total])
    return np.array(feats)

def frequency_ratio_features(window):
    feats = []
    for ch in range(window.shape[1]):
        freqs, psd = welch(window[:, ch], fs=fs, nperseg=fs)
        d, t, a, b = (band_power(freqs, psd, 0.5, 4), band_power(freqs, psd, 4, 8),
                      band_power(freqs, psd, 8, 13),   band_power(freqs, psd, 13, 30))
        feats.extend([t/b if b>0 else 0, t/a if a>0 else 0, a/b if b>0 else 0,
                      d/t if t>0 else 0, (t+a)/b if b>0 else 0])
    return np.array(feats)

def spectral_entropy_features(window):
    feats = []
    for ch in range(window.shape[1]):
        freqs, psd = welch(window[:, ch], fs=fs, nperseg=fs)
        psd_n = (psd + 1e-12) / np.sum(psd + 1e-12)
        feats.append(-np.sum(psd_n * np.log2(psd_n)))
    return np.array(feats)

def dfa_features(window):
    vals = []
    for ch in range(window.shape[1]):
        try:    vals.append(ant.detrended_fluctuation(window[:, ch]))
        except: vals.append(0)
    return np.array(vals)

def hurst_features(window):
    vals = []
    for ch in range(window.shape[1]):
        try:    vals.append(ant.hurst_rs(window[:, ch]))
        except: vals.append(0.5)
    return np.array(vals)

def fooof_features(window):
    feats = []
    fm = FOOOF(peak_width_limits=(1,8), max_n_peaks=5, aperiodic_mode='fixed', verbose=False)
    for ch in range(window.shape[1]):
        freqs, psd = welch(window[:, ch], fs=fs, nperseg=fs)
        try:
            fm.fit(freqs, psd)
            off, exp = fm.aperiodic_params_
            pk = fm.get_params('peak_params')
            al = pk[(pk[:,0]>=8) & (pk[:,0]<=13)]
            feats.extend([off, exp, al[0,0], al[0,1]] if len(al)>0 else [off, exp, 0, 0])
        except:
            feats.extend([0, 0, 0, 0])
    return np.array(feats)

def advanced_features(window):
    feats = []
    for ch in range(window.shape[1]):
        x = window[:, ch]
        try:
            coeffs = pywt.wavedec(x, 'db4', level=4)
            for c in coeffs: feats.append(np.sum(c**2))
        except:
            feats.extend([0]*5)
        try:    feats.append(ant.lziv_complexity(x))
        except: feats.append(0)
    return np.array(feats)

def coherence_features(window):
    """
    Spectral coherence between all pairs of channels.
    19 channels → 171 pairs × 4 bands = 684 features.
    """
    n_ch  = window.shape[1]
    bands = [(1, 4), (4, 8), (8, 13), (13, 30)]
    feats = []
    for i in range(n_ch):
        for j in range(i + 1, n_ch):
            f, Cxy = coherence(window[:, i], window[:, j], fs=fs, nperseg=fs)
            for fmin, fmax in bands:
                idx = (f >= fmin) & (f <= fmax)
                feats.append(np.mean(Cxy[idx]) if np.any(idx) else 0.0)
    return np.array(feats)


def extract_window_features(window):
    return np.concatenate([
        time_features(window), frequency_power_features(window),
        frequency_ratio_features(window), spectral_entropy_features(window),
        dfa_features(window), hurst_features(window),
        fooof_features(window), advanced_features(window),
        coherence_features(window)
    ])

# =======================================================
# STATISTICS MODE
# =======================================================
def Calc_stats(X_subj):
    stat_funcs = {
        "mean": np.mean, "std": np.std, "median": np.median,
        "iqr": scipy.stats.iqr, "skew": scipy.stats.skew, "kurtosis": scipy.stats.kurtosis,
    }
    stats_mat = np.zeros((X_subj.shape[1], len(prms["ftrs_stats"])))
    for i in range(X_subj.shape[1]):
        for j, stat in enumerate(prms["ftrs_stats"]):
            stats_mat[i, j] = stat_funcs[stat](X_subj[:, i])
    return stats_mat.flatten(order="C")

# =======================================================
# FEATURE SELECTION
# =======================================================
def features_scoring(feature_values, labels):
    best_acc = 0
    for th in np.unique(feature_values):
        best_acc = max(best_acc,
                       np.mean((feature_values > th).astype(int) == labels),
                       np.mean((feature_values <= th).astype(int) == labels))
    return best_acc

def _mrmr_suffix(n_selected):
    return f"pure_mrmr_top{n_selected}_{ANALYSIS_MODE}_w{prms['window_sec']}_o{int(prms['overlap_pct']*100)}.npy"

def rank_features_mrmr(X, y, n_selected=200):
    full_ranking_path = os.path.join(save_folder, _mrmr_suffix(n_selected))

    if os.path.exists(full_ranking_path):
        print(f"Loading full PURE MRMR ranking from cache ({full_ranking_path})")
        full_ranking = np.load(full_ranking_path).tolist()
    else:
        n_features  = X.shape[1]
        max_to_rank = n_selected if n_selected is not None else n_features
        print(f"Running PURE MRMR. Ranking top {max_to_rank} out of {n_features} features. Please wait...")
        relevance = mutual_info_classif(X, y, random_state=42, n_jobs=1)
        selected  = [int(np.argmax(relevance))]
        remaining = list(set(range(n_features)) - set(selected))

        def score_candidate(cand):
            rel = relevance[cand]
            red = np.mean([mutual_info_regression(X[:, cand].reshape(-1,1), X[:, s])[0] for s in selected])
            return rel - red, cand

        for step in range(max_to_rank - 1):
            results = Parallel(n_jobs=N_JOBS)(delayed(score_candidate)(c) for c in remaining)
            best    = max(results)[1]
            selected.append(best)
            remaining.remove(best)
            if (step + 1) % 10 == 0:
                print(f"    Ranked {step+1}/{max_to_rank} features...")

        full_ranking = selected
        np.save(full_ranking_path, np.array(full_ranking))
        print(f"Pure MRMR ranking saved to: {full_ranking_path}")

    return np.array(full_ranking)

# =======================================================
# SUBJECT LEVEL SPLIT
# =======================================================
def subject_level_split(X_all, y_all, subj_ids, test_pct=0.2, random_state=42):
    rng        = np.random.default_rng(random_state)
    unique_ids = np.unique(subj_ids)
    adhd_ids   = [s for s in unique_ids if s in ADHD_IDS]
    ctrl_ids   = [s for s in unique_ids if s in CONTROL_IDS]
    test_subjs = (set(rng.choice(adhd_ids, max(1, int(len(adhd_ids)*test_pct)), replace=False)) |
                  set(rng.choice(ctrl_ids,  max(1, int(len(ctrl_ids)*test_pct)),  replace=False)))
    train_mask = np.array([s not in test_subjs for s in subj_ids])
    return X_all[train_mask], X_all[~train_mask], y_all[train_mask], y_all[~train_mask]

# =======================================================
# EXCEL OUTPUT
# ← שינוי: עמודת "Test Accuracy" נוספה ל-CV Results
# =======================================================
def save_results_to_excel(cv_log, test_log, run_params, path):
    header_fill = PatternFill("solid", start_color="1F4E79")
    header_font = Font(bold=True, color="FFFFFF", name="Arial", size=11)
    param_fill  = PatternFill("solid", start_color="D9E1F2")
    best_fill   = PatternFill("solid", start_color="E2EFDA")
    center      = Alignment(horizontal="center", vertical="center")
    thin        = Side(style="thin", color="BFBFBF")
    border      = Border(left=thin, right=thin, top=thin, bottom=thin)

    param_cols = [
        ("Timestamp",     run_params["timestamp"]),
        ("Analysis Mode", run_params["analysis_mode"]),
        ("Selection Mode",run_params["selection_mode"]),
        ("N Features",    run_params["n_selected"]),
        ("Window Sec",    run_params["window_sec"]),
        ("Overlap %",     run_params["overlap_pct"]),
        ("CV Folds",      run_params["n_folds"]),
        ("Test Size %",   run_params["test_size_pct"]),
        ("Train Samples", run_params["train_samples"]),
        ("Test Samples",  run_params["test_samples"]),
    ]
    param_names = [x[0] for x in param_cols]
    param_vals  = [x[1] for x in param_cols]

    # מילון מהיר: שם מודל → Test Accuracy (לשימוש בלשונית CV)
    test_acc_by_model = {r["Model"]: round(r["Accuracy"], 4) for r in test_log} if test_log else {}

    def style_cell(cell, fill=None):
        cell.alignment = center
        cell.border    = border
        if fill: cell.fill = fill

    file_exists = os.path.exists(path)
    wb = load_workbook(path) if file_exists else Workbook()
    if not file_exists and "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    # ---------------------------------------------------
    # CV Results — כולל עמודת Test Accuracy להשוואה
    # ---------------------------------------------------
    cv_headers = param_names + ["Model", "CV Mean", "CV Std", "Test Accuracy"]  # ← נוסף
    if "CV Results" not in wb.sheetnames:
        ws1 = wb.create_sheet("CV Results")
        for col, h in enumerate(cv_headers, 1):
            cell = ws1.cell(row=1, column=col, value=h)
            cell.fill, cell.font, cell.alignment, cell.border = header_fill, header_font, center, border
    else:
        ws1 = wb["CV Results"]

    if cv_log:
        best_cv = max(x["CV Mean"] for x in cv_log)
        row = ws1.max_row + 1
        for r in cv_log:
            fill     = best_fill if r["CV Mean"] == best_cv else None
            test_acc = test_acc_by_model.get(r["Model"], "")   # ← שליפת Test Accuracy לפי מודל
            vals     = param_vals + [r["Model"], round(r["CV Mean"], 4), round(r["CV Std"], 4), test_acc]
            for col, val in enumerate(vals, 1):
                style_cell(ws1.cell(row=row, column=col, value=val),
                           fill if col > len(param_vals) else param_fill)
            row += 1

    # ---------------------------------------------------
    # Test Results — לשונית נפרדת, ללא שינוי
    # ---------------------------------------------------
    if "Test Results" not in wb.sheetnames:
        ws2 = wb.create_sheet("Test Results")
        for col, h in enumerate(param_names + ["Model","Accuracy","F1","AUC","TP","FP","FN","TN"], 1):
            cell = ws2.cell(row=1, column=col, value=h)
            cell.fill, cell.font, cell.alignment, cell.border = header_fill, header_font, center, border
    else:
        ws2 = wb["Test Results"]

    if test_log:
        best_test = max(x["Accuracy"] for x in test_log)
        row = ws2.max_row + 1
        for r in test_log:
            fill = best_fill if r["Accuracy"] == best_test else None
            vals = param_vals + [r["Model"], round(r["Accuracy"],4), round(r["F1"],4),
                                 round(r["AUC"],4), r["TP"], r["FP"], r["FN"], r["TN"]]
            for col, val in enumerate(vals, 1):
                style_cell(ws2.cell(row=row, column=col, value=val),
                           fill if col > len(param_vals) else param_fill)
            row += 1

    wb.save(path)
    print(f"Excel updated: {path}")

# =======================================================
# PREPROCESSING WORKER
# =======================================================
def process_subject(folder_path, i, label):
    file_path = os.path.join(folder_path, f"{i}.mat")
    if not os.path.exists(file_path):
        return
    mat      = loadmat(file_path)
    raw_data = mat[[k for k in mat if not k.startswith('__')][0]]
    data_f   = highpass_filter(lowpass_filter(raw_data))
    if prms["apply_zscore"]:
        data_f = zscore_normalization(data_f)
    windows = segment_signal(data_f)
    feats   = np.array([extract_window_features(w) for w in windows])
    np.save(os.path.join(save_folder, f"features_subj{i}.npy"), feats)
    np.save(os.path.join(save_folder, f"labels_subj{i}.npy"), np.full(len(feats), label))
    print(f"Subject {i} done ({len(windows)} windows)")

# =======================================================
# SINGLE RUN FUNCTION
# =======================================================
def run_single(n_selected, X_all, y_all, subj_ids, precomputed_ranking,
               X_train, X_test, y_train, y_test, skf):
    prms["n_selected"] = n_selected
    print(f"\n{'='*60}")
    print(f"=== MODE: {ANALYSIS_MODE} | SELECTION: {FEATURE_SELECTION_MODE} "
          f"| n_selected={n_selected} | folds={prms['n_folds']} ===")
    print(f"{'='*60}\n")

    selected_idx = precomputed_ranking[:n_selected]
    X_train_sel  = X_train[:, selected_idx]
    X_test_sel   = X_test[:,  selected_idx]

    if prms["apply_final_zscore"]:
        mean = X_train_sel.mean(axis=0)
        std  = X_train_sel.std(axis=0)
        std[std == 0] = 1
        X_train_sel = (X_train_sel - mean) / std
        X_test_sel  = (X_test_sel  - mean) / std

    ts       = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    cv_log   = []
    test_log = []

    if RUN_MODEL_TOURNAMENT:
        print("\n--- CROSS VALIDATION (StratifiedKFold) ---")
        def run_cv(m):
            scores = cross_val_score(
                m["class"](**m["params"]),
                X_train_sel, y_train,
                cv=skf,
                n_jobs=1,
            )
            return {"Model": m["name"], "CV Mean": float(np.mean(scores)), "CV Std": float(np.std(scores))}
        cv_log = Parallel(n_jobs=N_JOBS)(delayed(run_cv)(m) for m in MODELS_TO_COMPARE)
        for r in cv_log:
            print(f"  {r['Model']:16s} | CV = {r['CV Mean']:.4f} ± {r['CV Std']:.4f}")

    if RUN_FINAL_TEST:
        print("\n--- FINAL TEST ---")
        for m in MODELS_TO_COMPARE:
            clf    = m["class"](**m["params"])
            clf.fit(X_train_sel, y_train)
            y_pred = clf.predict(X_test_sel)
            acc    = accuracy_score(y_test, y_pred)
            f1     = f1_score(y_test, y_pred, zero_division=0)
            try:    auc = roc_auc_score(y_test, clf.predict_proba(X_test_sel)[:, 1])
            except: auc = np.nan
            tn, fp, fn, tp = confusion_matrix(y_test, y_pred).ravel()
            test_log.append({"Model": m["name"], "Accuracy": float(acc), "F1": float(f1),
                             "AUC": float(auc), "TP": int(tp), "FP": int(fp), "FN": int(fn), "TN": int(tn)})
            print(f"  {m['name']:16s} | ACC={acc:.4f} | F1={f1:.4f}")

    if cv_log or test_log:
        save_results_to_excel(cv_log, test_log, {
            "timestamp":     ts,
            "analysis_mode": ANALYSIS_MODE,
            "selection_mode":FEATURE_SELECTION_MODE,
            "n_selected":    n_selected,
            "window_sec":    prms["window_sec"],
            "overlap_pct":   int(prms["overlap_pct"] * 100),
            "n_folds":       prms["n_folds"],
            "test_size_pct": int(prms["test_size_pct"] * 100),
            "train_samples": len(y_train),
            "test_samples":  len(y_test),
        }, excel_log_path)

# =======================================================
# MAIN
# =======================================================
if __name__ == "__main__":

    if RUN_PREPROCESSING:
        print("--- PHASE 1: PREPROCESSING ---")
        jobs = [(fp, i, 0 if i in ADHD_IDS else 1) for fp, ids in folders.items() for i in ids]
        Parallel(n_jobs=N_JOBS)(delayed(process_subject)(fp, i, lbl) for fp, i, lbl in jobs)

    if RUN_BUILD_DATASET:
        print("--- PHASE 2: BUILD DATASET ---")
        X_list, y_list, s_list = [], [], []
        for i in range(1, 122):
            fp = os.path.join(save_folder, f"features_subj{i}.npy")
            lp = os.path.join(save_folder, f"labels_subj{i}.npy")
            if not os.path.exists(fp):
                continue
            feats = np.load(fp)
            label = np.load(lp)[0]
            if ANALYSIS_MODE == "windowed":
                X_list.append(feats)
                y_list.append(np.full(len(feats), label))
                s_list.append(np.full(len(feats), i))
            elif ANALYSIS_MODE == "statistics":
                X_list.append(Calc_stats(feats).reshape(1, -1))
                y_list.append([label])
                s_list.append([i])

        X_all    = np.vstack(X_list)
        y_all    = np.concatenate(y_list)
        subj_ids = np.concatenate(s_list)
    else:
        raise RuntimeError("RUN_BUILD_DATASET=False — הוסף טעינה ידנית של X_all/y_all/subj_ids")

    print(f"Dataset: {X_all.shape} | ADHD={np.sum(y_all==0)} | Control={np.sum(y_all==1)}")

    X_train, X_test, y_train, y_test = subject_level_split(
        X_all, y_all, subj_ids, test_pct=prms["test_size_pct"]
    )
    print(f"Train: {len(y_train)} | Test: {len(y_test)}")

    print("Cleaning NaNs/Infs from Train and Test sets...")
    for dist_set, name in zip([X_train, X_test], ["Train", "Test"]):
        dist_set[np.isinf(dist_set)] = np.nan
        all_nan_cols = np.isnan(dist_set).all(axis=0)
        if np.any(all_nan_cols):
            dist_set[:, all_nan_cols] = 0.0
        nan_counts = np.isnan(dist_set).sum(axis=0)
        bad_cols   = np.where(nan_counts > 0)[0]
        if len(bad_cols) > 0:
            print(f"  [{name} Set] Found {len(bad_cols)} features containing NaNs. Fixing with column mean...")
            for col in bad_cols:
                col_mean = np.nanmean(dist_set[:, col])
                if np.isnan(col_mean): col_mean = 0.0
                dist_set[np.isnan(dist_set[:, col]), col] = col_mean
    print("Data splits are clean and ready!")

    print(f"\n--- LOADING RANKING ({FEATURE_SELECTION_MODE}) ---")
    precomputed_ranking = rank_features_mrmr(X_train, y_train, n_selected=200)
    print(f"Ranking ready: {len(precomputed_ranking)} features ranked.\n")

    skf = StratifiedKFold(n_splits=prms["n_folds"], shuffle=True, random_state=42)
    print(f"Running with n_folds={prms['n_folds']}")

    for n_sel in range(3, 201):
        run_single(n_sel, X_all, y_all, subj_ids,
                   precomputed_ranking, X_train, X_test, y_train, y_test, skf)

    print("\n=== ALL RUNS COMPLETE ===\n")